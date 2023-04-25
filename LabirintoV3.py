import random
import openpyxl
from pyamaze import maze, agent, COLOR,textLabel

def DFS(m, start=None, goal=None):
    # Define o ponto de início da busca como sendo o ponto (rows, cols) do labirinto
    if start is None:
        start = (m.rows, m.cols)
        
    # Define o ponto final da busca como sendo o ponto _goal do labirinto
    if goal is None:
        goal = m._goal

    # Inicializa a lista de pontos já explorados com o ponto inicial
    explored = [start]
    # Inicializa a lista de pontos a serem explorados com o ponto inicial
    frontier = [start]
    # Inicializa o dicionário que irá armazenar o caminho percorrido
    path = {}
    # Inicializa a lista de nós explorados
    nodes = []

    # Enquanto a lista de pontos a serem explorados não estiver vazia
    while len(frontier) > 0:
        # Remove o último elemento da lista de pontos a serem explorados
        node = frontier.pop()
        # Adiciona o nó à lista de nós explorados
        nodes.append(node)
        
        # Se o nó atual for o objetivo da busca, encerra o loop
        if node == goal:
            break
            
        # Cria uma lista vazia para os filhos do nó atual
        children = []
        # Para cada direção possível
        for d in 'ESNW':
            # Se a direção for possível a partir do nó atual
            if m.maze_map[node][d] == True:
                # Define o próximo nó a ser explorado naquela direção
                if d == 'E':
                    child = (node[0], node[1]+1)
                elif d == 'W':
                    child = (node[0], node[1]-1)
                elif d == 'N':
                    child = (node[0]-1, node[1])
                else:
                    child = (node[0]+1, node[1])
                
                # Se o filho já foi explorado, passa para o próximo filho
                if child in explored:
                    continue
                
                # Adiciona o filho à lista de filhos
                children.append(child)
                # Adiciona o filho à lista de pontos explorados
                explored.append(child)
                # Adiciona o filho à lista de pontos a serem explorados
                frontier.append(child)
                
                # Armazena o caminho percorrido
                path[child] = node
                
        # Se o nó atual tiver mais de um filho, marca o nó atual como uma célula de marcação
        if len(children) > 1:
            m.markCells.append(node)

    # Inicializa o dicionário que irá armazenar o caminho percorrido mais curto
    shortest_path = {}
    # Define o ponto atual como sendo o objetivo
    current = goal

    # Enquanto o ponto atual não for o ponto de início da busca
    while current != start:
        # Adiciona o ponto atual ao caminho percorrido mais curto
        shortest_path[path[current]] = current
        # Define o ponto atual como sendo o pai do ponto atual
        current = path[current]
    
    # Retorna as listas e dicionários com as informações obtidas pela busca em profundidade
    return nodes, path, shortest_path


if __name__=='__main__':
   
    
    # Solicita a largura e altura do labirinto ao usuário por meio do terminal e converte as entradas em inteiros
    LARGURA_LABIRINTO = int(input("Digite a largura do labirinto: "))
    ALTURA_LABIRINTO = int(input("Digite a altura do labirinto: "))

    nome_janela = "Meu Labirinto"
        
    # Cria uma instância do labirinto com a largura e altura fornecidas
    m=maze(LARGURA_LABIRINTO,ALTURA_LABIRINTO) 

    
    IniCom = random.randint(1, LARGURA_LABIRINTO)
    IniLar = random.randint(1, ALTURA_LABIRINTO)
    FimCom = random.randint(1, LARGURA_LABIRINTO)
    FimLar = random.randint(1, LARGURA_LABIRINTO)
    # Define o ponto de partida e o objetivo do labirinto
    start = (IniCom, IniLar)
    goal = (FimCom, FimLar)
    
    
    # Cria um labirinto com o algoritmo "Recursive Backtracker"
    m.CreateMaze(IniCom,IniLar) 

    # Executa o algoritmo de busca em profundidade no labirinto para encontrar o caminho do ponto de partida até o objetivo
    nodes, path, shortest_path = DFS(m, start, goal) 

    # Cria uma instância de um agente que percorre o caminho mais curto do ponto de partida até o objetivo
    a=agent(m, start[0], start[1], goal=goal, footprints=True, shape='square', color=COLOR.green)

    # Cria uma instância de um agente que mostra o caminho completo do ponto de partida até o objetivo
    b=agent(m, goal[0], goal[1], goal=start, footprints=True, filled=True)

    # Mostra o caminho completo do ponto de partida até o objetivo com os nós percorridos marcados em verde
    m.tracePath({a: nodes}, showMarked=True)

    # Mostra apenas o caminho percorrido do ponto de partida até o objetivo
    m.tracePath({b: path})

    l=textLabel(m,'Passos',len(path)+1)
    # In=textLabel(m,IniCom,IniLar)
    # Fn=textLabel(m,FimCom, FimLar)
    
    # Abre o arquivo de planilha do Excel (se existir) ou cria um novo arquivo se ele não existir
    workbook = openpyxl.load_workbook("labirinto.xlsx")
   
    worksheet = workbook.active

    # Escreve os títulos das colunas
    worksheet.cell(row=1, column=1).value = "Altura"
    worksheet.cell(row=1, column=2).value = "Largura"
    worksheet.cell(row=1, column=3).value = "Ponto Inicial"
    worksheet.cell(row=1, column=4).value = "Ponto Final"
    worksheet.cell(row=1, column=5).value = "Quantidade de Passos"

    # Escreve os dados do labirinto na planilha do Excel
    row = worksheet.max_row + 1
    worksheet.cell(row=row, column=1).value = ALTURA_LABIRINTO
    worksheet.cell(row=row, column=2).value = LARGURA_LABIRINTO
    worksheet.cell(row=row, column=3).value = f"({IniCom}, {IniLar})"
    worksheet.cell(row=row, column=4).value = f"({FimCom}, {FimLar})"
    worksheet.cell(row=row, column=5).value = len(path)+1

    # Salva o arquivo de planilha do Excel
    workbook.save("labirinto.xlsx")

    # Executa a simulação do labirinto com o agente percorrendo o caminho mais curto do ponto de partida até o objetivo
    m.run()


