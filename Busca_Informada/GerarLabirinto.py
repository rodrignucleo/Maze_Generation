import random
from pyamaze import maze, agent, textLabel, COLOR
import openpyxl

from PosicoesLabirinto import PosicoesLabirinto

def GerarLabirinto(LARGURA_LABIRINTO, ALTURA_LABIRINTO):
    labirinto = maze(LARGURA_LABIRINTO, ALTURA_LABIRINTO)

    # Procura um numero randomicamente 
    RANDOM_X_COMECO = random.randint(1, LARGURA_LABIRINTO-1)
    RANDOM_Y_COMECO = random.randint(1, ALTURA_LABIRINTO-1)

    RANDOM_X_FIM = random.randint(1, LARGURA_LABIRINTO)
    RANDOM_Y_FIM = random.randint(1, ALTURA_LABIRINTO)

    labirinto.CreateMaze(RANDOM_X_FIM, RANDOM_Y_FIM, loopPercent=10)

    path = PosicoesLabirinto(labirinto, RANDOM_X_COMECO, RANDOM_Y_COMECO, RANDOM_X_FIM, RANDOM_Y_FIM)

    # Cria um objeto para mostrar o trajeto/passos feitos
    passos = agent(labirinto, x=RANDOM_X_COMECO, y=RANDOM_Y_COMECO, shape='arrow', footprints=True)
    labirinto.tracePath({passos: path})
    

    # Uma caixa de mensagem com a quantidade de passos dados
    l = textLabel(labirinto, 'Passos dados: ', len(path))

    # Abre o arquivo de planilha do Excel (se existir)
    workbook = openpyxl.load_workbook("labirinto_star_algoritmo.xlsx")

    # workbook = openpyxl.Workbook()
    worksheet = workbook.active

    # Escreve os t√≠tulos das colunas
    worksheet.cell(row=1, column=1).value = "Altura"
    worksheet.cell(row=1, column=2).value = "Largura"
    worksheet.cell(row=1, column=3).value = "Ponto Inicial"
    worksheet.cell(row=1, column=4).value = "Ponto Final"
    worksheet.cell(row=1, column=5).value = "Quantidade de Passos"

    # Escreve os dados do labirinto na planilha do Excel
    row = worksheet.max_row + 1
    worksheet.cell(row=row, column=1).value = ALTURA_LABIRINTO
    worksheet.cell(row=row, column=2).value = LARGURA_LABIRINTO
    worksheet.cell(row=row, column=3).value = f"({RANDOM_X_COMECO}, {RANDOM_Y_COMECO})"
    worksheet.cell(row=row, column=4).value = f"({RANDOM_X_FIM}, {RANDOM_Y_FIM})"
    worksheet.cell(row=row, column=5).value = len(path)

    # Salva o arquivo de planilha do Excel
    workbook.save("labirinto_star_algoritmo.xlsx")

    labirinto.run()